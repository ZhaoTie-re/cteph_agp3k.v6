#!/usr/bin/env python3

import argparse
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="Build a PLINK update table with FID IID SEX PHENO columns from an Excel sheet."
	)
	parser.add_argument("--xlsx", required=True, help="Input Excel workbook path")
	parser.add_argument("--out", required=True, help="Output TSV path")
	parser.add_argument("--sample-id-col", required=True, help="Column name containing sample IDs")
	parser.add_argument("--pheno-col", required=True, help="Column name containing phenotype labels")
	parser.add_argument("--case-value", required=True, help="Phenotype value to encode as case (2)")
	parser.add_argument("--ctrl-value", required=True, help="Phenotype value to encode as ctrl (1)")
	parser.add_argument("--sex-col", required=True, help="Column name containing sex labels")
	parser.add_argument("--female-value", required=True, help="Sex value to encode as female (2)")
	parser.add_argument("--male-value", required=True, help="Sex value to encode as male (1)")
	parser.add_argument("--fam", help="Optional PLINK .fam file to validate sample coverage against")
	return parser.parse_args()


def normalize(value) -> str | None:
	if value is None:
		return None
	text = str(value).strip()
	return text if text else None


def load_records(args: argparse.Namespace) -> dict[str, tuple[str, str]]:
	workbook = load_workbook(args.xlsx, read_only=True, data_only=True)
	worksheet = workbook[workbook.sheetnames[0]]
	header = [normalize(cell) for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
	col_index = {name: idx for idx, name in enumerate(header) if name is not None}

	required = [args.sample_id_col, args.pheno_col, args.sex_col]
	missing_cols = [col for col in required if col not in col_index]
	if missing_cols:
		raise ValueError(f"Missing required column(s): {', '.join(missing_cols)}")

	case_value = normalize(args.case_value)
	ctrl_value = normalize(args.ctrl_value)
	female_value = normalize(args.female_value)
	male_value = normalize(args.male_value)

	records: dict[str, tuple[str, str]] = {}
	duplicate_ids: set[str] = set()

	for row in worksheet.iter_rows(min_row=2, values_only=True):
		sample_id = normalize(row[col_index[args.sample_id_col]])
		pheno_raw = normalize(row[col_index[args.pheno_col]])
		sex_raw = normalize(row[col_index[args.sex_col]])

		if sample_id is None or pheno_raw is None or sex_raw is None:
			continue

		if pheno_raw == case_value:
			pheno = "2"
		elif pheno_raw == ctrl_value:
			pheno = "1"
		else:
			continue

		if sex_raw == male_value:
			sex = "1"
		elif sex_raw == female_value:
			sex = "2"
		else:
			continue

		record = (sex, pheno)
		if sample_id in records and records[sample_id] != record:
			duplicate_ids.add(sample_id)
			continue
		records[sample_id] = record

	if duplicate_ids:
		raise ValueError(
			"Conflicting duplicate sample IDs found in sample_info: " + ", ".join(sorted(duplicate_ids)[:10])
		)

	return records


def write_table(records: dict[str, tuple[str, str]], out_path: Path) -> None:
	with out_path.open("w", encoding="utf-8", newline="") as handle:
		writer = csv.writer(handle, delimiter="\t")
		writer.writerow(["#FID", "IID", "SEX", "PHENO"])
		for sample_id in sorted(records):
			sex, pheno = records[sample_id]
			writer.writerow([sample_id, sample_id, sex, pheno])


def validate_fam(records: dict[str, tuple[str, str]], fam_path: Path) -> tuple[int, int]:
	fam_ids: list[str] = []
	with fam_path.open("r", encoding="utf-8") as handle:
		for line in handle:
			parts = line.rstrip("\n").split()
			if len(parts) < 2:
				continue
			fam_ids.append(parts[1])

	missing = [sample_id for sample_id in fam_ids if sample_id not in records]
	if missing:
		preview = ", ".join(missing[:10])
		raise ValueError(
			f"{len(missing)} sample(s) in {fam_path} are missing phenotype/sex annotations. Example IDs: {preview}"
		)

	return len(fam_ids), len(records)


def main() -> int:
	args = parse_args()
	try:
		records = load_records(args)
		out_path = Path(args.out)
		write_table(records, out_path)

		message = f"Wrote {len(records)} update records to {out_path}"
		if args.fam:
			fam_count, update_count = validate_fam(records, Path(args.fam))
			message += f"; validated {fam_count} PLINK samples against {update_count} annotated records"
		print(message, file=sys.stderr)
		return 0
	except Exception as exc:  # noqa: BLE001
		print(f"ERROR: {exc}", file=sys.stderr)
		return 1


if __name__ == "__main__":
	sys.exit(main())